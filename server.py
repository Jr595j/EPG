"""
server.py — Flask web server for the EPG Aggregator.

Endpoints:
  GET /                  — Status dashboard (HTML)
  GET /epg.xml           — Merged XMLTV file (point TiviMate here)
  GET /refresh           — Force a manual re-fetch and redirect to dashboard
  GET /status            — Machine-readable JSON status
  GET /channels          — Searchable list of all EPG channel IDs
  GET /preview           — Search a channel and see its programme schedule
  GET /guide             — Visual TV guide grid (like a cable box EPG)
  GET /suggest-mappings  — Download Excel with fuzzy-matched EPG↔M3U suggestions
"""

import io
import logging
import os
from datetime import datetime, timezone
from html import escape

try:
    from defusedxml.ElementTree import parse as _safe_parse, fromstring as _safe_fromstring
    from xml.etree import ElementTree as ET
    ET.parse = _safe_parse          # type: ignore[assignment]
    ET.fromstring = _safe_fromstring  # type: ignore[assignment]
except ImportError:
    from xml.etree import ElementTree as ET

from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, jsonify, redirect, send_file, Response, request

import fetcher

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
scheduler = BackgroundScheduler(daemon=True)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

STYLE = """
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', monospace; background: #0f0f1a; color: #dde; max-width: 860px; margin: 40px auto; padding: 20px; }
  h1 { color: #00d4ff; font-size: 1.6em; margin-bottom: 4px; }
  h2 { color: #00d4ff; font-size: 1.1em; border-bottom: 1px solid #223; padding-bottom: 6px; margin: 24px 0 10px; }
  .subtitle { color: #667; font-size: 0.85em; margin-bottom: 20px; }
  .card { background: #16213e; border-radius: 8px; padding: 16px 20px; margin: 10px 0; }
  .stats { display: flex; gap: 32px; }
  .stat-val { font-size: 2.2em; color: #00d4ff; font-weight: bold; }
  .stat-lbl { font-size: 0.75em; color: #667; text-transform: uppercase; letter-spacing: .06em; }
  .badge { display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 0.78em; font-weight: bold; }
  .badge-ok { background: #1a4a2a; color: #4caf50; }
  .badge-err { background: #4a1a1a; color: #f44336; }
  .badge-off { background: #2a2a3a; color: #667; }
  a { color: #00d4ff; text-decoration: none; }
  a:hover { text-decoration: underline; }
  .btn { display: inline-block; background: #00d4ff; color: #000; padding: 7px 18px; border-radius: 5px; margin: 4px 4px 4px 0; font-weight: bold; font-size: 0.9em; }
  .btn:hover { background: #00b8e0; text-decoration: none; }
  .btn-sec { background: #1e3a5f; color: #00d4ff; }
  .btn-sec:hover { background: #254a78; }
  table { width: 100%; border-collapse: collapse; font-size: 0.9em; }
  td, th { padding: 7px 10px; border-bottom: 1px solid #1e2a40; text-align: left; }
  th { color: #667; font-size: 0.8em; text-transform: uppercase; letter-spacing: .06em; }
  code { background: #0d1525; padding: 2px 7px; border-radius: 4px; color: #00d4ff; font-size: 0.95em; }
  input[type=text] { width: 100%; background: #0d1525; border: 1px solid #223; color: #dde; padding: 8px 12px; border-radius: 5px; font-size: 0.95em; margin-bottom: 12px; }
  .ts { color: #667; font-size: 0.85em; }
  .err-box { background: #2a1010; border: 1px solid #5a2020; border-radius: 6px; padding: 10px 14px; color: #f88; font-size: 0.85em; margin-top: 8px; }
</style>
"""


def _page(title: str, body: str, wide: bool = False) -> str:
    extra = "<style>body{max-width:none;}</style>" if wide else ""
    return f"<!DOCTYPE html><html><head><meta charset='utf-8'><title>{title}</title>{STYLE}{extra}</head><body>{body}</body></html>"


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    config = fetcher.load_config()
    r = fetcher.last_result
    instance_name = config.get("instance_name", "EPG Aggregator")

    # Status badge
    if not r:
        badge = "<span class='badge badge-off'>Not yet fetched</span>"
    elif r.get("status") == "ok":
        badge = "<span class='badge badge-ok'>OK</span>"
    else:
        badge = "<span class='badge badge-err'>Error</span>"

    # Stats row
    channels = r.get("channels", "—")
    programmes = r.get("programmes", "—")
    ts = r.get("timestamp", "—")

    stats_html = f"""
    <div class='stats'>
      <div><div class='stat-val'>{channels}</div><div class='stat-lbl'>Channels</div></div>
      <div><div class='stat-val'>{programmes}</div><div class='stat-lbl'>Programmes</div></div>
      <div><div class='stat-val'>{r.get('sources_ok','—')}</div><div class='stat-lbl'>Sources OK</div></div>
    </div>
    <br><span class='ts'>Last refresh: {ts} UTC &nbsp; {badge}</span>
    """

    errors_html = ""
    for err in r.get("errors", []):
        errors_html += f"<div class='err-box'>⚠ {escape(err['source'])}: {escape(err['error'])}</div>"

    # Sources table
    source_rows = ""
    for s in config.get("sources", []):
        enabled = s.get("enabled", True)
        status_badge = "<span class='badge badge-ok'>enabled</span>" if enabled else "<span class='badge badge-off'>disabled</span>"
        source_rows += f"<tr><td>{s['name']}</td><td>{s.get('priority',1)}</td><td>{status_badge}</td></tr>"

    port = config.get("port", 8080)
    refresh_h = config.get("refresh_interval_hours", 12)

    # Channel map summary
    ch_map = {k: v for k, v in config.get("channel_map", {}).items() if not k.startswith("_")}
    map_summary = f"{len(ch_map)} remapping(s) configured" if ch_map else "No remappings yet — <a href='/channels'>check /channels</a> to find IDs to map"

    body = f"""
    <h1>{instance_name}</h1>
    <p class='subtitle'>Serving merged XMLTV for TiviMate · Auto-refresh every {refresh_h}h</p>

    <div class='card'>
      {stats_html}
      {errors_html}
    </div>

    <div>
      <a class='btn' href='/epg.xml'>📺 EPG URL (copy for TiviMate)</a>
      <a class='btn btn-sec' href='/refresh'>↻ Force Refresh</a>
      <a class='btn btn-sec' href='/guide'>📺 Channel Guide</a>
      <a class='btn btn-sec' href='/channels'>🔍 Channel ID Browser</a>
      <a class='btn btn-sec' href='/preview'>📋 Programme Preview</a>
      <a class='btn btn-sec' href='/suggest-mappings'>📊 Suggest EPG Mappings</a>
      <a class='btn btn-sec' href='/status'>{ '{}' } JSON Status</a>
    </div>

    <h2>TiviMate Setup</h2>
    <div class='card'>
      In TiviMate: <em>Settings → Playlists → [your playlist] → EPG sources → Add</em><br><br>
      EPG URL: <code>http://YOUR-PC-IP:{port}/epg.xml</code><br><br>
      Replace <code>YOUR-PC-IP</code> with your PC's local IP address (e.g. 192.168.1.x).<br>
      Run <code>ipconfig</code> in Command Prompt to find it.
    </div>

    <h2>Sources ({len(config.get('sources', []))})</h2>
    <div class='card'>
      <table><thead><tr><th>Name</th><th>Priority</th><th>Status</th></tr></thead>
      <tbody>{source_rows}</tbody></table>
    </div>

    <h2>Channel ID Mapping</h2>
    <div class='card'>
      {map_summary}<br><br>
      Edit <code>config.json</code> → <code>channel_map</code> to remap EPG channel IDs so they match your M3U <code>tvg-id</code> values.
    </div>
    """
    return _page("EPG Aggregator", body)


@app.route("/epg.xml")
def serve_epg():
    if not os.path.exists(fetcher.MERGED_PATH):
        return Response("EPG not yet generated. Visit /refresh first.", status=503, mimetype="text/plain")
    return send_file(fetcher.MERGED_PATH, mimetype="application/xml")


@app.route("/refresh")
def force_refresh():
    logger.info("Manual refresh triggered via /refresh")
    _do_fetch()
    return redirect("/")


@app.route("/status")
def status():
    return jsonify(fetcher.last_result)


@app.route("/channels")
def channels_page():
    channels = fetcher.list_channels()
    if not channels:
        return _page("Channel IDs", "<div class='card'>EPG not yet generated. <a href='/refresh'>Refresh now</a>.</div>")

    rows = "".join(
        f"<tr><td>{escape(c['name'])}</td><td><code>{escape(c['id'])}</code></td></tr>"
        for c in channels
    )

    body = f"""
    <h1><a href='/'>← EPG Aggregator</a></h1>
    <h2>EPG Channel IDs &nbsp;<span style='color:#667;font-size:.8em;font-weight:normal'>({len(channels)} total)</span></h2>
    <div class='card'>
      <p style='margin-bottom:12px;color:#aab;font-size:.9em'>
        Your M3U <code>tvg-id</code> values must exactly match the <strong>ID</strong> column below.<br>
        If they don't match, add entries to <code>channel_map</code> in <code>config.json</code>:<br>
        <code>"epg-id-shown-here": "tvg-id-from-your-m3u"</code>
      </p>
      <input type='text' id='q' placeholder='Filter by name or ID...' oninput='filter()'>
      <table>
        <thead><tr><th>Display Name</th><th>EPG ID (must match M3U tvg-id)</th></tr></thead>
        <tbody id='tbody'>{rows}</tbody>
      </table>
    </div>
    <script>
      function filter() {{
        var q = document.getElementById('q').value.toLowerCase();
        document.querySelectorAll('#tbody tr').forEach(function(r) {{
          r.style.display = r.textContent.toLowerCase().indexOf(q) >= 0 ? '' : 'none';
        }});
      }}
    </script>
    """
    return _page("EPG Channel IDs", body)


@app.route("/preview")
def preview():
    if not os.path.exists(fetcher.MERGED_PATH):
        return _page("Programme Preview", "<div class='card'>EPG not yet generated. <a href='/refresh'>Refresh now</a>.</div>")

    query = request.args.get("q", "").strip().lower()

    # Parse merged XML
    tree = ET.parse(fetcher.MERGED_PATH)
    root = tree.getroot()

    # Build channel id → display name map
    ch_names = {}
    for ch in root.findall("channel"):
        ch_id = ch.get("id", "")
        name_el = ch.find("display-name")
        ch_names[ch_id] = (name_el.text or ch_id) if name_el is not None else ch_id

    results_html = ""

    if query:
        # Find matching channel IDs
        matched_ids = [
            ch_id for ch_id, name in ch_names.items()
            if query in name.lower() or query in ch_id.lower()
        ]
        matched_ids_set = set(matched_ids)

        if not matched_ids:
            results_html = f"<div class='card' style='color:#f88'>No channels found matching <strong>{escape(query)}</strong>.</div>"
        else:
            # Collect programmes for matched channels
            ch_programmes = {ch_id: [] for ch_id in matched_ids}
            for prog in root.findall("programme"):
                ch_id = prog.get("channel", "")
                if ch_id not in matched_ids_set:
                    continue
                start_raw = prog.get("start", "")
                title_el  = prog.find("title")
                desc_el   = prog.find("desc")
                title = (title_el.text or "") if title_el is not None else ""
                desc  = (desc_el.text  or "") if desc_el  is not None else ""

                # Parse timestamp with proper timezone handling
                dt = fetcher.parse_xmltv_time(start_raw)
                if dt is not None:
                    start_fmt = dt.strftime("%a %b %d  %H:%M")
                else:
                    start_fmt = start_raw

                ch_programmes[ch_id].append({
                    "start": start_fmt,
                    "start_raw": start_raw,
                    "title": title,
                    "desc": desc[:120] + ("…" if len(desc) > 120 else ""),
                })

            for ch_id in sorted(matched_ids, key=lambda x: ch_names[x].lower()):
                progs = sorted(ch_programmes[ch_id], key=lambda p: (
                    fetcher.parse_xmltv_time(p["start_raw"]) or datetime.min.replace(tzinfo=timezone.utc),
                    p["start_raw"]
                ))
                name = ch_names[ch_id]
                count = len(progs)

                if count == 0:
                    prog_rows = "<tr><td colspan='3' style='color:#f88'>⚠ Channel exists in EPG but has NO programme data — source does not cover this channel</td></tr>"
                    status_badge = "<span class='badge badge-err'>No programmes</span>"
                else:
                    earliest = progs[0]["start"]
                    latest   = progs[-1]["start"]
                    status_badge = f"<span class='badge badge-ok'>{count} programmes · {earliest} → {latest}</span>"
                    # Highlight currently airing / upcoming
                    now_utc = datetime.now(timezone.utc)
                    now_str = now_utc.strftime("%Y%m%d%H%M%S")
                    prog_rows = ""
                    for p in progs[:48]:
                        # Compare normalized UTC start time for highlight
                        norm_start = fetcher.parse_xmltv_time(p["start_raw"])
                        if norm_start:
                            is_future = norm_start.strftime("%Y%m%d%H%M%S") >= now_str
                        else:
                            is_future = p["start_raw"][:14] >= now_str
                        row_style = " style='background:#0d1f0d'" if is_future else ""
                        prog_rows += (
                            f"<tr{row_style}>"
                            f"<td style='white-space:nowrap;color:#aab'>{p['start']}</td>"
                            f"<td><strong>{escape(p['title'])}</strong></td>"
                            f"<td style='color:#889;font-size:.85em'>{escape(p['desc'])}</td></tr>"
                        )
                    if count > 48:
                        prog_rows += f"<tr><td colspan='3' style='color:#667'>… and {count - 48} more</td></tr>"

                results_html += f"""
                <div class='card' style='margin-bottom:16px'>
                  <div style='margin-bottom:8px'>
                    <strong style='color:#00d4ff'>{escape(name)}</strong> &nbsp;
                    <code style='font-size:.8em'>{escape(ch_id)}</code> &nbsp;
                    {status_badge}
                  </div>
                  <table>
                    <thead><tr><th>Time (UTC)</th><th>Programme</th><th>Description</th></tr></thead>
                    <tbody>{prog_rows}</tbody>
                  </table>
                </div>"""

    body = f"""
    <h1><a href='/'>← EPG Aggregator</a></h1>
    <h2>Programme Preview</h2>
    <div class='card'>
      <p style='color:#aab;font-size:.9em;margin-bottom:10px'>
        Search a channel to see its programme data in the merged EPG.<br>
        <strong style='color:#4caf50'>Data shown here = data TiviMate receives.</strong>
        If data is here but TiviMate shows blank, the issue is the <code>tvg-id</code> mismatch — check <a href='/channels'>/channels</a>.
        If data is missing here, the source isn't covering that channel.
      </p>
      <form method='get' action='/preview' style='display:flex;gap:8px'>
        <input type='text' name='q' value='{escape(query)}' placeholder='Search channel name e.g. HBO, ESPN, CNN...' style='margin:0;flex:1'>
        <button type='submit' style='background:#00d4ff;color:#000;border:none;padding:8px 18px;border-radius:5px;font-weight:bold;cursor:pointer'>Search</button>
      </form>
    </div>
    {results_html}
    """
    return _page("Programme Preview", body)


@app.route("/guide")
def guide():
    """Visual TV guide grid — channels × time, like a cable box EPG."""
    if not os.path.exists(fetcher.MERGED_PATH):
        return _page("Channel Guide", "<div class='card'>EPG not yet generated. <a href='/refresh'>Refresh now</a>.</div>")

    from datetime import timedelta

    query = request.args.get("q", "").strip().lower()
    # Time window: default 4h starting from current half-hour
    now = datetime.now(timezone.utc)
    # Snap to previous half-hour
    grid_start = now.replace(minute=(now.minute // 30) * 30, second=0, microsecond=0)
    try:
        hours = int(request.args.get("hours", 4))
    except (ValueError, TypeError):
        hours = 4
    hours = max(1, min(hours, 24))
    grid_end = grid_start + timedelta(hours=hours)

    # Pixels per minute for the grid
    px_per_min = 4
    total_width = hours * 60 * px_per_min

    tree = ET.parse(fetcher.MERGED_PATH)
    root = tree.getroot()

    # Build channel list with names
    channels = []
    for ch in root.findall("channel"):
        ch_id = ch.get("id", "")
        name_el = ch.find("display-name")
        name = (name_el.text or ch_id) if name_el is not None else ch_id
        channels.append((ch_id, name))

    # Index programmes by channel
    ch_progs = {}
    for prog in root.findall("programme"):
        ch_id = prog.get("channel", "")
        start_dt = fetcher.parse_xmltv_time(prog.get("start", ""))
        stop_dt = fetcher.parse_xmltv_time(prog.get("stop", ""))
        if not start_dt or not stop_dt:
            continue
        # Only include programmes that overlap the grid window
        if stop_dt <= grid_start or start_dt >= grid_end:
            continue
        title_el = prog.find("title")
        desc_el = prog.find("desc")
        title = (title_el.text or "") if title_el is not None else ""
        desc = (desc_el.text or "") if desc_el is not None else ""
        ch_progs.setdefault(ch_id, []).append({
            "start": start_dt,
            "stop": stop_dt,
            "title": title,
            "desc": desc,
        })

    # Filter to channels that have programmes (or match query)
    filtered = []
    for ch_id, name in channels:
        has_progs = ch_id in ch_progs
        if query:
            if query not in name.lower() and query not in ch_id.lower():
                continue
        else:
            if not has_progs:
                continue
        filtered.append((ch_id, name))

    filtered.sort(key=lambda c: c[1].lower())

    # Build time header slots (every 30 min)
    time_headers = ""
    t = grid_start
    slot_width = 30 * px_per_min
    while t < grid_end:
        label = t.strftime("%H:%M")
        time_headers += f"<div class='g-time' style='width:{slot_width}px'>{label}</div>"
        t += timedelta(minutes=30)

    # Build channel rows
    rows_html = ""
    for ch_id, name in filtered:
        progs = sorted(ch_progs.get(ch_id, []), key=lambda p: p["start"])
        blocks = ""
        for p in progs:
            # Clamp to grid window
            vis_start = max(p["start"], grid_start)
            vis_stop = min(p["stop"], grid_end)
            offset_min = (vis_start - grid_start).total_seconds() / 60
            dur_min = (vis_stop - vis_start).total_seconds() / 60
            if dur_min < 1:
                continue
            left = offset_min * px_per_min
            width = dur_min * px_per_min - 2  # 2px gap
            is_now = p["start"] <= now < p["stop"]
            cls = "g-prog g-now" if is_now else "g-prog"
            t_start = p["start"].strftime("%H:%M")
            t_stop = p["stop"].strftime("%H:%M")
            safe_title = escape(p["title"])
            safe_desc = escape(p["desc"][:200])
            blocks += (
                f"<div class='{cls}' style='left:{left:.0f}px;width:{width:.0f}px' "
                f"title='{t_start}-{t_stop}  {safe_title}\n{safe_desc}'>"
                f"<span class='g-title'>{safe_title}</span>"
                f"<span class='g-time-sm'>{t_start}</span>"
                f"</div>"
            )
        safe_name = escape(name)
        rows_html += (
            f"<div class='g-row'>"
            f"<div class='g-ch'>{safe_name}</div>"
            f"<div class='g-progs' style='width:{total_width}px'>{blocks}</div>"
            f"</div>"
        )

    guide_style = f"""
    <style>
      .guide-wrap {{ max-width: none; }}
      .g-controls {{ display:flex; gap:8px; align-items:center; flex-wrap:wrap; margin-bottom:12px; }}
      .g-controls input[type=text] {{ width:260px; margin:0; }}
      .g-controls select, .g-controls button {{ background:#0d1525; border:1px solid #223; color:#dde; padding:6px 12px; border-radius:5px; font-size:.9em; }}
      .g-controls button {{ background:#00d4ff; color:#000; font-weight:bold; cursor:pointer; border:none; }}
      .g-container {{ overflow-x:auto; border:1px solid #1e2a40; border-radius:8px; background:#0a0a18; }}
      .g-header {{ display:flex; position:sticky; top:0; z-index:2; background:#16213e; }}
      .g-header .g-ch-hdr {{ width:180px; min-width:180px; padding:8px 10px; font-size:.75em; color:#667;
                             text-transform:uppercase; letter-spacing:.06em; border-right:1px solid #1e2a40;
                             position:sticky; left:0; background:#16213e; z-index:3; }}
      .g-time {{ padding:8px 0; text-align:center; font-size:.75em; color:#667; border-left:1px solid #1e2a40; }}
      .g-row {{ display:flex; border-top:1px solid #111; }}
      .g-ch {{ width:180px; min-width:180px; padding:6px 10px; font-size:.82em; color:#aab;
               white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
               border-right:1px solid #1e2a40; background:#0f0f1a;
               position:sticky; left:0; z-index:1; }}
      .g-progs {{ position:relative; height:42px; }}
      .g-prog {{ position:absolute; top:2px; height:38px; background:#1a3050; border-radius:4px;
                 overflow:hidden; padding:3px 6px; cursor:default; border:1px solid #234; }}
      .g-prog:hover {{ background:#1e4070; z-index:5; }}
      .g-now {{ background:#1a4a2a; border-color:#2a6a3a; }}
      .g-now:hover {{ background:#1e5a30; }}
      .g-title {{ font-size:.78em; color:#dde; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; display:block; }}
      .g-time-sm {{ font-size:.65em; color:#667; }}
      .g-nowline {{ position:absolute; top:0; bottom:0; width:2px; background:#f44; z-index:4; pointer-events:none; }}
    </style>
    """

    now_offset = (now - grid_start).total_seconds() / 60 * px_per_min
    now_line = f"<div class='g-nowline' style='left:{180 + now_offset:.0f}px'></div>" if 0 <= now_offset <= total_width else ""

    grid_start_str = grid_start.strftime("%H:%M")
    grid_end_str = grid_end.strftime("%H:%M")

    body = f"""
    {guide_style}
    <div class='guide-wrap'>
    <h1><a href='/'>← EPG Aggregator</a></h1>
    <h2>Channel Guide &nbsp;<span style='color:#667;font-size:.8em;font-weight:normal'>{grid_start_str}–{grid_end_str} UTC · {len(filtered)} channels</span></h2>
    <div class='card'>
      <form method='get' action='/guide' class='g-controls'>
        <input type='text' name='q' value='{escape(query)}' placeholder='Filter channels (e.g. HBO, ESPN, Fox...)'>
        <select name='hours'>
          {"".join(f"<option value='{h}'{' selected' if h == hours else ''}>{h}h</option>" for h in [2, 4, 6, 8, 12, 24])}
        </select>
        <button type='submit'>Update</button>
      </form>
      <p style='color:#667;font-size:.8em;margin-top:4px'>
        <span style='display:inline-block;width:12px;height:12px;background:#1a4a2a;border-radius:2px;vertical-align:middle'></span> Currently airing &nbsp;
        <span style='display:inline-block;width:12px;height:2px;background:#f44;vertical-align:middle'></span> Now &nbsp;
        Hover a programme for details. Times are UTC.
      </p>
    </div>
    <div class='g-container' style='position:relative'>
      {now_line}
      <div class='g-header'>
        <div class='g-ch-hdr'>Channel</div>
        {time_headers}
      </div>
      {rows_html}
      {f"<div class='card' style='margin:20px;color:#667'>No channels found{' matching ' + chr(34) + escape(query) + chr(34) if query else ''}.</div>" if not filtered else ""}
    </div>
    </div>
    """
    return _page("Channel Guide", body, wide=True)


@app.route("/suggest-mappings")
def suggest_mappings():
    """
    Download an Excel workbook that fuzzy-matches your M3U tvg-ids against
    the merged EPG channel IDs and suggests entries for channel_map.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    config = fetcher.load_config()
    m3u_url = config.get("m3u_url", "").strip()

    if not m3u_url:
        return Response("m3u_url not set in config.json", status=400, mimetype="text/plain")
    if not os.path.exists(fetcher.MERGED_PATH):
        return Response("EPG not yet generated — visit /refresh first.", status=503, mimetype="text/plain")

    # ── 1. Fetch M3U ────────────────────────────────────────────────────────
    try:
        m3u_channels = fetcher.parse_m3u(m3u_url)
    except Exception as exc:
        return Response(f"Failed to fetch M3U: {exc}", status=502, mimetype="text/plain")

    # ── 2. Load EPG channels from merged cache ───────────────────────────────
    tree = ET.parse(fetcher.MERGED_PATH)
    epg_root = tree.getroot()

    epg_channels = []
    for ch in epg_root.findall("channel"):
        epg_id = ch.get("id", "")
        name_el = ch.find("display-name")
        epg_name = (name_el.text or "") if name_el is not None else ""
        epg_channels.append((epg_id, epg_name))

    epg_id_set = {c[0] for c in epg_channels}

    # Normalized lookups: norm_key → (epg_id, epg_name)
    norm_id_map   = {}
    norm_name_map = {}
    for epg_id, epg_name in epg_channels:
        nid   = fetcher.normalize_name(epg_id)
        nname = fetcher.normalize_name(epg_name)
        if nid   and nid   not in norm_id_map:   norm_id_map[nid]   = (epg_id, epg_name)
        if nname and nname not in norm_name_map: norm_name_map[nname] = (epg_id, epg_name)

    # Fuzzy corpus: list of (normalized_key, epg_id, epg_name)
    fuzzy_corpus = []
    seen_keys = set()
    for epg_id, epg_name in epg_channels:
        for raw in (epg_id, epg_name):
            nk = fetcher.normalize_name(raw)
            if nk and nk not in seen_keys:
                seen_keys.add(nk)
                fuzzy_corpus.append((nk, epg_id, epg_name))
    fuzzy_keys = [f[0] for f in fuzzy_corpus]

    try:
        from rapidfuzz import fuzz
        from rapidfuzz import process as rfprocess
        has_rapidfuzz = True
    except ImportError:
        has_rapidfuzz = False

    def find_match(tvg_id, tvg_name, display_name):
        """Returns (match_type, epg_id, epg_name, score)."""
        # 1. Exact
        if tvg_id and tvg_id in epg_id_set:
            epg_name = next((n for i, n in epg_channels if i == tvg_id), "")
            return ("Exact", tvg_id, epg_name, 100)

        # 2. Normalized (try all three M3U name fields)
        for query in (tvg_id, tvg_name, display_name):
            if not query:
                continue
            nq = fetcher.normalize_name(query)
            if not nq:
                continue
            if nq in norm_id_map:
                return ("Normalized", *norm_id_map[nq], 95)
            if nq in norm_name_map:
                return ("Normalized", *norm_name_map[nq], 92)

        # 3. Fuzzy
        if has_rapidfuzz and fuzzy_keys:
            best_score, best_idx = 0, 0
            for query in (tvg_id, tvg_name, display_name):
                if not query:
                    continue
                nq = fetcher.normalize_name(query)
                if not nq or len(nq) < 3:
                    continue
                result = rfprocess.extractOne(nq, fuzzy_keys, scorer=fuzz.token_sort_ratio)
                if result and result[1] > best_score:
                    best_score, best_idx = result[1], result[2]
            if best_score >= 70:
                _, epg_id, epg_name = fuzzy_corpus[best_idx]
                return ("Fuzzy", epg_id, epg_name, best_score)

        return ("No Match", "", "", 0)

    # ── 3. Run matching ──────────────────────────────────────────────────────
    results = []
    for ch in m3u_channels:
        tvg_id       = ch["tvg_id"]
        tvg_name     = ch["tvg_name"]
        display_name = ch["name"]
        group        = ch["group"]

        match_type, epg_id, epg_name, score = find_match(tvg_id, tvg_name, display_name)

        map_entry = ""
        if match_type not in ("Exact", "No Match") and epg_id and tvg_id:
            map_entry = f'"{epg_id}": "{tvg_id}"'

        results.append({
            "group":      group,
            "name":       display_name,
            "tvg_id":     tvg_id,
            "match_type": match_type,
            "score":      score,
            "epg_id":     epg_id,
            "epg_name":   epg_name,
            "map_entry":  map_entry,
        })

    # Sort: No Match → Fuzzy → Normalized → Exact; then group → name
    _order = {"No Match": 0, "Fuzzy": 1, "Normalized": 2, "Exact": 3}
    results.sort(key=lambda r: (_order.get(r["match_type"], 9), r["group"].lower(), r["name"].lower()))

    # ── 4. Build Excel ───────────────────────────────────────────────────────
    wb = Workbook()

    # ── Sheet 1: All channels ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Mapping Suggestions"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["#", "Group", "M3U Name", "M3U tvg-id", "Status", "Score", "EPG ID", "EPG Display Name", "Paste into channel_map"]
    col_widths = [5,   22,      36,          36,            14,       7,       36,        36,                  48]

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(results) + 1}"

    row_fills = {
        "Exact":      PatternFill("solid", start_color="E2EFDA"),  # green
        "Normalized": PatternFill("solid", start_color="DDEBF7"),  # blue
        "Fuzzy":      PatternFill("solid", start_color="FFF2CC"),  # yellow
        "No Match":   PatternFill("solid", start_color="FCE4D6"),  # red
    }
    data_font = Font(name="Arial", size=10)
    mono_font = Font(name="Courier New", size=9)

    for i, r in enumerate(results, 1):
        row  = i + 1
        fill = row_fills.get(r["match_type"], row_fills["No Match"])
        score_display = r["score"] if r["match_type"] == "Fuzzy" else ""

        vals = [i, r["group"], r["name"], r["tvg_id"], r["match_type"],
                score_display, r["epg_id"], r["epg_name"], r["map_entry"]]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font      = mono_font if col == 9 else data_font

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total      = len(results)
    n_exact    = sum(1 for r in results if r["match_type"] == "Exact")
    n_norm     = sum(1 for r in results if r["match_type"] == "Normalized")
    n_fuzzy    = sum(1 for r in results if r["match_type"] == "Fuzzy")
    n_none     = sum(1 for r in results if r["match_type"] == "No Match")
    fill_now   = f"{n_exact / total * 100:.1f}%" if total else "0%"
    fill_pot   = f"{(n_exact + n_norm + n_fuzzy) / total * 100:.1f}%" if total else "0%"

    summary_rows = [
        ["EPG Mapping Analysis", ""],
        ["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
        ["", ""],
        ["Total M3U channels",                              total],
        ["Exact Match  (already working in TiviMate)",      n_exact],
        ["Normalized Match  (add to channel_map)",          n_norm],
        ["Fuzzy Match  (review, then add to channel_map)",  n_fuzzy],
        ["No Match Found",                                  n_none],
        ["", ""],
        ["Current fill rate",                               fill_now],
        ["Potential fill rate (with all suggestions)",      fill_pot],
    ]

    for row_data in summary_rows:
        ws2.append(row_data)

    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 20
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws2["A2"].font = Font(name="Arial", size=10, color="888888")
    ws2["B2"].font = Font(name="Arial", size=10, color="888888")

    lbl_font = Font(name="Arial", size=11)
    val_font = Font(name="Arial", bold=True, size=11)
    for row_idx in range(4, 12):
        ws2.cell(row=row_idx, column=1).font = lbl_font
        ws2.cell(row=row_idx, column=2).font = val_font

    ws2["B5"].fill  = row_fills["Exact"]
    ws2["B6"].fill  = row_fills["Normalized"]
    ws2["B7"].fill  = row_fills["Fuzzy"]
    ws2["B8"].fill  = row_fills["No Match"]

    # ── Serialize and return ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="epg_mapping_suggestions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Scheduler
# ---------------------------------------------------------------------------

def _do_fetch():
    try:
        fetcher.run_fetch()
    except Exception as exc:
        logger.error(f"Fetch failed: {exc}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="EPG Aggregator Server")
    parser.add_argument("--config", default="config.json",
                        help="Path to config JSON file (default: config.json)")
    args = parser.parse_args()

    # Point fetcher at the chosen config before loading anything
    fetcher.set_config_path(args.config)

    config = fetcher.load_config()
    port = config.get("port", 8080)
    refresh_hours = config.get("refresh_interval_hours", 12)
    instance_name = config.get("instance_name", "EPG Aggregator")

    logger.info("=" * 60)
    logger.info(f"{instance_name} starting...")
    logger.info(f"  Config    : {args.config}")
    logger.info(f"  Dashboard : http://localhost:{port}/")
    logger.info(f"  EPG URL   : http://localhost:{port}/epg.xml")
    logger.info(f"  Refresh   : every {refresh_hours} hour(s)")
    logger.info("=" * 60)

    # Fetch immediately on startup
    _do_fetch()

    # Schedule recurring fetch
    scheduler.add_job(_do_fetch, "interval", hours=refresh_hours, id="epg_refresh")
    scheduler.start()

    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
